import xml.etree.ElementTree as ET
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# 1. Parse the XML file
xml_file = r"C:\Users\Lenovo\Desktop\CreditLink_Response\response.xml"
tree = ET.parse(xml_file)
root = tree.getroot()

# Define MISMO & XLINK namespaces
ns = {
    'mismo': 'http://www.mismo.org/residential/2009/schemas',
    'xlink': 'http://www.w3.org/1999/xlink'
}

data_tables = {}

# ==========================================
# BACKGROUND MAP: DETECT XLINK RELATIONSHIPS
# ==========================================
borrower_map = {}
for party in root.findall('.//mismo:PARTY', ns):
    party_id = party.get(f"{{{ns['xlink']}}}label")
    fullname_elem = party.find('.//mismo:INDIVIDUAL/mismo:NAME/mismo:FullName', ns)
    if party_id and fullname_elem is not None:
        borrower_map[party_id] = fullname_elem.text

score_to_borrower_map = {}
for link in root.findall('.//mismo:CREDIT_SCORE_LINK', ns):
    score_ref = link.get(f"{{{ns['xlink']}}}href")   
    party_ref = link.get(f"{{{ns['xlink']}}}from")   
    if score_ref and party_ref:
        score_to_borrower_map[score_ref.replace('#', '')] = party_ref

for rel in root.findall('.//mismo:RELATIONSHIP', ns):
    from_ref = rel.get(f"{{{ns['xlink']}}}from")
    to_ref = rel.get(f"{{{ns['xlink']}}}to")
    if from_ref and to_ref:
        clean_from = from_ref.replace('#', '')
        clean_to = to_ref.replace('#', '')
        if clean_from in borrower_map:
            score_to_borrower_map[clean_to] = clean_from
        elif clean_to in borrower_map:
            score_to_borrower_map[clean_from] = clean_to


# ==========================================
# 1. LOAN SUMMARY DATA
# ==========================================
loan_data = []
for loan in root.findall('.//mismo:LOAN', ns):
    loan_id = loan.find('.//mismo:LOAN_IDENTIFIER/mismo:LoanIdentifier', ns)
    sys_id = loan.find('.//mismo:ORIGINATION_SYSTEM/mismo:LoanOriginationSystemLoanIdentifier', ns)
    sys_name = loan.find('.//mismo:ORIGINATION_SYSTEM/mismo:LoanOriginationSystemName', ns)
    purpose = loan.find('.//mismo:TERMS_OF_LOAN/mismo:LoanPurposeType', ns)
    mortgage = loan.find('.//mismo:TERMS_OF_LOAN/mismo:MortgageType', ns)
    state_date = loan.find('.//mismo:LOAN_STATE/mismo:LoanStateDate', ns)
    state_type = loan.find('.//mismo:LOAN_STATE/mismo:LoanStateType', ns)
    
    loan_data.append({
        "LoanRoleType": loan.get("LoanRoleType", "N/A"),
        "LoanIdentifier": loan_id.text if loan_id is not None else "N/A",
        "LoanOriginationSystemLoanIdentifier": sys_id.text.strip() if sys_id is not None else "N/A",
        "LoanOriginationSystemName": sys_name.text.strip() if sys_name is not None else "N/A",
        "LoanPurposeType": purpose.text if purpose is not None else "N/A",
        "MortgageType": mortgage.text if mortgage is not None else "N/A",
        "LoanStateDate": state_date.text if state_date is not None else "N/A",
        "LoanStateType": state_type.text if state_type is not None else "N/A"
    })
data_tables["Loan Summary"] = pd.DataFrame(loan_data)


# ==========================================
# 2. BORROWERS & PARTIES
# ==========================================
party_data = []
for party in root.findall('.//mismo:PARTIES/mismo:PARTY', ns):
    ind = party.find('mismo:INDIVIDUAL', ns)
    legal = party.find('mismo:LEGAL_ENTITY', ns)
    tax_id = party.find('.//mismo:TAXPAYER_IDENTIFIER/mismo:TaxpayerIdentifierValue', ns)
    role = party.find('.//mismo:ROLE_DETAIL/mismo:PartyRoleType', ns)
    
    name, age, dob, marital, address_str = "N/A", "N/A", "N/A", "N/A", "N/A"
        
    if ind is not None:
        name_elem = ind.find('.//mismo:NAME/mismo:FullName', ns)
        name = name_elem.text if name_elem is not None else "N/A"
        
        detail = party.find('.//mismo:BORROWER_DETAIL', ns)
        if detail is not None:
            age = getattr(detail.find('mismo:BorrowerAgeAtApplicationYearsCount', ns), 'text', 'N/A')
            dob = getattr(detail.find('mismo:BorrowerBirthDate', ns), 'text', 'N/A')
            marital = getattr(detail.find('mismo:MaritalStatusType', ns), 'text', 'N/A')
            
        addr = party.find('.//mismo:RESIDENCES/mismo:RESIDENCE/mismo:ADDRESS', ns)
        if addr is not None:
            line = getattr(addr.find('mismo:AddressLineText', ns), 'text', '')
            city = getattr(addr.find('mismo:CityName', ns), 'text', '')
            st = getattr(addr.find('mismo:StateCode', ns), 'text', '')
            zipc = getattr(addr.find('mismo:PostalCode', ns), 'text', '')
            address_str = f"{line}, {city}, {st} {zipc}".strip(", ")

    if role is not None or name != "N/A":
        party_data.append({
            "FullName": name,
            "PartyRoleType": role.text if role is not None else "N/A",
            "TaxpayerIdentifierValue": tax_id.text if tax_id is not None else "N/A",
            "BorrowerAgeAtApplicationYearsCount": age,
            "BorrowerBirthDate": dob,
            "MaritalStatusType": marital,
            "AddressLineText": address_str if address_str else "N/A"
        })
data_tables["Parties & Borrowers"] = pd.DataFrame(party_data)


# ==========================================
# 3. CREDIT LIABILITIES
# ==========================================
liabilities_data = []
for liab in root.findall('.//mismo:CREDIT_LIABILITY', ns):
    seq = liab.get("SequenceNumber", "N/A")
    creditor = liab.find('.//mismo:CREDIT_LIABILITY_CREDITOR/mismo:NAME/mismo:FullName', ns)
    rating_type = liab.find('.//mismo:CREDIT_LIABILITY_CURRENT_RATING/mismo:CreditLiabilityCurrentRatingType', ns)
    
    detail = liab.find('mismo:CREDIT_LIABILITY_DETAIL', ns)
    acct_id, opened_dt, status_type, acct_type, loan_type = "N/A", "N/A", "N/A", "N/A", "N/A"
    bal_amt, pmnt_amt, limit_amt, past_due = 0, 0, 0, 0
    
    if detail is not None:
        acct_id = getattr(detail.find('mismo:CreditLiabilityAccountIdentifier', ns), 'text', 'N/A')
        opened_dt = getattr(detail.find('mismo:CreditLiabilityAccountOpenedDate', ns), 'text', 'N/A')
        status_type = getattr(detail.find('mismo:CreditLiabilityAccountStatusType', ns), 'text', 'N/A')
        acct_type = getattr(detail.find('mismo:CreditLiabilityAccountType', ns), 'text', 'N/A')
        loan_type = getattr(detail.find('mismo:CreditLoanType', ns), 'text', 'N/A')
        
        bal_amt = float(getattr(detail.find('mismo:CreditLiabilityUnpaidBalanceAmount', ns), 'text', 0) or 0)
        pmnt_amt = float(getattr(detail.find('mismo:CreditLiabilityMonthlyPaymentAmount', ns), 'text', 0) or 0)
        limit_amt = float(getattr(detail.find('mismo:CreditLiabilityCreditLimitAmount', ns), 'text', 0) or 0)
        past_due = float(getattr(detail.find('mismo:CreditLiabilityPastDueAmount', ns), 'text', 0) or 0)
        
    comment = liab.find('.//mismo:CREDIT_COMMENTS/mismo:CREDIT_COMMENT/mismo:CreditCommentText', ns)
    
    liabilities_data.append({
        "SequenceNumber": seq,
        "FullName": creditor.text if creditor is not None else "N/A",
        "CreditLiabilityAccountIdentifier": acct_id,
        "CreditLiabilityAccountType": acct_type,
        "CreditLoanType": loan_type,
        "CreditLiabilityAccountStatusType": status_type,
        "CreditLiabilityCurrentRatingType": rating_type.text if rating_type is not None else "N/A",
        "CreditLiabilityAccountOpenedDate": opened_dt,
        "CreditLiabilityUnpaidBalanceAmount": bal_amt,
        "CreditLiabilityMonthlyPaymentAmount": pmnt_amt,
        "CreditLiabilityCreditLimitAmount": limit_amt,
        "CreditLiabilityPastDueAmount": past_due,
        "CreditCommentText": comment.text if comment is not None else "N/A"
    })
data_tables["Credit Liabilities"] = pd.DataFrame(liabilities_data)


# ==========================================
# 4. CREDIT SCORES
# ==========================================
score_data = []
for score in root.findall('.//mismo:CREDIT_SCORE', ns):
    score_label = score.get(f"{{{ns['xlink']}}}label", "")
    linked_party_id = score_to_borrower_map.get(score_label)
    resolved_borrower_name = borrower_map.get(linked_party_id, "Unmapped/Joint Account")
    
    detail = score.find('mismo:CREDIT_SCORE_DETAIL', ns)
    score_value, model_name, provider_name = "N/A", "N/A", "N/A"
    
    if detail is not None:
        score_value_elem = detail.find('mismo:CreditScoreValue', ns)
        score_value = int(score_value_elem.text) if score_value_elem is not None and score_value_elem.text else "N/A"
        
        model_name_elem = detail.find('mismo:CreditScoreModelNameType', ns)
        model_name = model_name_elem.text if model_name_elem is not None else "N/A"
        
        provider_name_elem = detail.find('mismo:CreditScoreModelNameTypeOtherDescription', ns)
        provider_name = provider_name_elem.text if provider_name_elem is not None else "N/A"

    score_data.append({
        "FullName": resolved_borrower_name,
        "label": score_label,
        "CreditScoreValue": score_value,
        "CreditScoreModelNameType": model_name,
        "CreditScoreModelNameTypeOtherDescription": provider_name
    })
if score_data:
    data_tables["Credit Scores"] = pd.DataFrame(score_data)


# ==========================================
# 5. FIXED: CREDIT SUMMARIES (EXACT MAPPING)
# ==========================================
summary_rows = []
for summary in root.findall('.//mismo:CREDIT_SUMMARY', ns):
    summary_label = summary.get(f"{{{ns['xlink']}}}label", "")
    linked_party_id = score_to_borrower_map.get(summary_label)
    resolved_borrower_name = borrower_map.get(linked_party_id, "Subject Profile")
    
    # Loop through every single dataset block inside the summary
    for dataset in summary.findall('.//mismo:CREDIT_SUMMARY_DATA_SET', ns):
        seq_num = dataset.get("SequenceNumber", "N/A")
        name_elem = dataset.find('mismo:CreditSummaryDataSetName', ns)
        value_elem = dataset.find('mismo:CreditSummaryDataSetValue', ns)
        
        name_text = name_elem.text if name_elem is not None else "N/A"
        value_text = value_elem.text if value_elem is not None else "N/A"
        
        # Clean numeric values for formatting, keep percentages or strings clean
        processed_value = value_text
        if value_text and value_text.replace('.', '', 1).isdigit() and '%' not in value_text:
            processed_value = float(value_text) if '.' in value_text else int(value_text)
            
        summary_rows.append({
            "FullName": resolved_borrower_name,
            "label": summary_label,
            "SequenceNumber": seq_num,
            "CreditSummaryDataSetName": name_text,
            "CreditSummaryDataSetValue": processed_value
        })

if summary_rows:
    data_tables["Credit Summaries"] = pd.DataFrame(summary_rows)


# ==========================================
# 6. WORKBOOK COMPILATION & STYLING
# ==========================================
output_file = "Validated_MISMO_Export.xlsx"

with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
    for sheet_name, df in data_tables.items():
        df.to_excel(writer, sheet_name=sheet_name, index=False)
        
        workbook = writer.book
        worksheet = writer.sheets[sheet_name]
        
        header_font = Font(name='Consolas', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='2F5597', end_color='2F5597', fill_type='solid') 
        data_font = Font(name='Segoe UI', size=10, color='000000')
        
        center_align = Alignment(horizontal='center', vertical='center')
        left_align = Alignment(horizontal='left', vertical='center')
        right_align = Alignment(horizontal='right', vertical='center')
        
        thin_border = Border(
            left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'),
            top=Side(style='thin', color='D9D9D9'), bottom=Side(style='thin', color='D9D9D9')
        )
        
        # Style Headers
        for col_num, column_title in enumerate(df.columns, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = left_align
            cell.border = thin_border
        
        # Style Data Cells
        for row in range(2, worksheet.max_row + 1):
            name_cell_val = str(worksheet.cell(row=row, column=4).value or '') # Checks CreditSummaryDataSetName if in summary sheet
            
            for col in range(1, worksheet.max_column + 1):
                cell = worksheet.cell(row=row, column=col)
                cell.font = data_font
                cell.border = thin_border
                
                col_header = df.columns[col - 1]
                
                # Dynamic type matching based on headers and dataset contexts
                if col_header == "CreditSummaryDataSetValue":
                    if isinstance(cell.value, (int, float)) and any(x in name_cell_val for x in ["Balance", "Past Due", "Payment", "Limit", "High Credit"]):
                        cell.number_format = '$#,##0.00'
                        cell.alignment = right_align
                    elif isinstance(cell.value, (int, float)):
                        cell.alignment = center_align
                    else:
                        cell.alignment = center_align # percentages / dates
                elif col_header in ["CreditLiabilityUnpaidBalanceAmount", "CreditLiabilityMonthlyPaymentAmount", "CreditLiabilityCreditLimitAmount", "CreditLiabilityPastDueAmount"]:
                    cell.number_format = '$#,##0.00'
                    cell.alignment = right_align
                elif col_header in ["SequenceNumber", "BorrowerAgeAtApplicationYearsCount", "BorrowerBirthDate", "CreditLiabilityAccountOpenedDate", "LoanStateDate", "TaxpayerIdentifierValue", "CreditScoreValue", "label"]:
                    cell.alignment = center_align
                else:
                    cell.alignment = left_align
                    
        # Apply Auto-fit Column Widths
        for col in worksheet.columns:
            max_len = 0
            for cell in col:
                col_h = df.columns[cell.column-1]
                if col_h == "CreditSummaryDataSetValue" and cell.row > 1 and '$' in getattr(cell, 'number_format', ''):
                    val_str = f"${cell.value:,.2f}" if isinstance(cell.value, (int, float)) else str(cell.value or '')
                elif "Amount" in col_h and cell.row > 1 and isinstance(cell.value, (int, float)):
                    val_str = f"${cell.value:,.2f}"
                else:
                    val_str = str(cell.value or '')
                if len(val_str) > max_len:
                    max_len = len(val_str)
            col_letter = get_column_letter(col[0].column)
            worksheet.column_dimensions[col_letter].width = max(max_len + 4, 15)

print(f"Spreadsheet rewritten successfully with exact dataset schema rules. File: '{output_file}'")